
import openpyxl
from tkinter import *
from tkinter import messagebox
from datetime import datetime as dt
from win32com import client as win32
from datetime import datetime
import os


window = Tk()

absolute_path = Label(text="path:")
absolute_path.grid(row=1, column=0)
absolute_path_entry = Entry(width=35)
absolute_path_entry.grid(row=1, column=1, columnspan=2)
absolute_path_entry.focus()

name_label = Label(text="Enter your name:")
name_label.grid(row=2, column=0)
name_entry = Entry(width=35)
name_entry.grid(row=2, column=1, columnspan=2)









# make sure to pass in your workbook
#request_num,origin,customer,name,total_requests
def generate_email():
    name = name_entry.get()
    file_path = absolute_path_entry.get()
    workbook = openpyxl.load_workbook(file_path)
    sh = workbook.active
    list_of_data = []
    
    #min_row=1, min_col=1, max_row=10, max_col=16
    list_of_rmots = []
    for row in sh.iter_rows(): 
        for cell in row:  
            if cell.column  == 1:
                list_of_rmots.append(cell.value)
    list_of_data.append(list_of_rmots)
    list_of_originators = []
    for row in sh.iter_rows(): 
        for cell in row: 
            if cell.column  == 15:
                list_of_originators.append(cell.value)
    list_of_data.append(list_of_originators)
    list_of_customer = []
    for row in sh.iter_rows(): 
        for cell in row: 
            if cell.column  == 2:
                list_of_customer.append(cell.value)
    list_of_data.append(list_of_customer)
    
    
    data = list_of_data
        
    rmots = data[0]
    origins = data[1]
    customers = data[2]
    total_requests = len(rmots)


    for n in range(0,total_requests):
        today = dt.now()
        time = today.time()
        hour = time.strftime("%H")

        if int(hour) < 12:
            greeting = "Good morning"
        elif int(hour) >= 12:
            greeting = "Good afternoon"


        outlook = win32.Dispatch('outlook.application')
        mail = outlook.CreateItem(0)
        mail.Subject = 'Engineer for request ' + rmots[n]
        mail.To = origins[n] + "@microsoft.com"
    


        email_text = f'''{greeting},<br><br>
        Did you still require an engineer for {customers[n]}?<br><br>
        Very Respectfully,<br><br>
        {name}'''


        mail.HTMLBody = email_text
        
        mail.Send()



'''
column 1 is the RMOT or tracing number
column 3 is the offering
column 12 is the start date format is dd-mmm-yyyy
column 8 is the country
column 15 is the originator
'''
#BUTTONS  command needs to be lambda, this allows you to pass arguments into your function button.
#without it python will just execute the function
send_emails = Button(text="Send Automated Emails", command=generate_email)
send_emails.grid(row=3, column=2)


window.mainloop()